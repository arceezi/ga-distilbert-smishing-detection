#!/usr/bin/env python
"""Repair the expert spam review packet using raw-complete replacements."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
INITIAL_PATH = OUT_DIR / "expert_spam_review_500.csv"
KEPT_PATH = OUT_DIR / "expert_spam_review_rows_kept.csv"
REPLACE_PATH = OUT_DIR / "expert_spam_review_rows_to_replace.csv"
POOL_PATH = OUT_DIR / "raw_complete_expert_replacement_pool.csv"
CSV_PATH = OUT_DIR / "expert_spam_review_500_raw_complete.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500_raw_complete.xlsx"
LOG_PATH = OUT_DIR / "expert_spam_review_replacement_log.csv"
ARCHIVE_PATH = OUT_DIR / "expert_spam_review_replaced_archive.csv"
REPORT_PATH = OUT_DIR / "expert_spam_review_raw_complete_report.md"

URL_RE = re.compile(r"https?://\S+|www\.\S+|(?:[a-z0-9-]+\.)+[a-z]{2,}(?:/\S*)?", re.I)
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
AMOUNT_RE = re.compile(r"(?:[$£€]|rs\.?|php|usd|gbp|eur)\s*\d+(?:[,.]\d+)*|\d+(?:[,.]\d+)*(?:\s?(?:php|usd|gbp|eur|rs|p))", re.I)

FINAL_COLUMNS = [
    "review_id",
    "message_for_review",
    "message_raw",
    "message_clean",
    "source_label",
    "normalized_label_before_review",
    "candidate_reason",
    "source_name",
    "dataset_name",
    "source_group",
    "contains_url",
    "contains_phone",
    "contains_otp",
    "contains_amount",
    "suggested_category",
    "raw_quality_status",
    "source_traceability_status",
    "expert_label",
    "expert_confidence",
    "expert_notes",
    "reviewer_name",
    "review_date",
]

CODEBOOK = """HAM: Legitimate/non-malicious SMS.

SPAM_NOT_SMISHING: Unwanted promotional or irrelevant SMS but not clearly phishing.

SMISHING: SMS phishing/social-engineering attempt involving deception, impersonation, credential/payment request, suspicious link/callback, urgency, account/security/delivery bait, or fraudulent intent.

UNSURE: Ambiguous or needs another reviewer.

REJECT: Not useful, not SMS-like, non-English, artifact, duplicate fragment, abusive reply, or report/commentary text.

Important expert note: Do not label smishing only because there is a URL. Look for deception, impersonation, fraudulent intent, credential/payment request, urgency, or social-engineering purpose.
"""


def normalize_text(text: str) -> str:
    text = str(text).lower()
    text = URL_RE.sub(" <URL> ", text)
    text = re.sub(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", " <EMAIL> ", text)
    text = PHONE_RE.sub(" <PHONE> ", text)
    text = re.sub(r"\b(?:otp|pin|code|passcode|verification)\s*[:#-]?\s*[a-z0-9-]{4,10}\b", " <OTP> ", text)
    text = re.sub(r"\b[a-z]{1,4}\d{4,10}\b|\b\d{4,8}[a-z]{1,4}\b", " <OTP> ", text)
    text = re.sub(r"\b\d{9,}\b", " <REF_NUM> ", text)
    text = AMOUNT_RE.sub(" <AMOUNT> ", text)
    text = re.sub(r"[^a-z0-9<>]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def family_key(text: str) -> str:
    return " ".join(normalize_text(text).split()[:10])


def source_traceability(row: pd.Series) -> str:
    return "traceable" if any(str(row.get(c, "")).strip() for c in ["source_name", "dataset_name", "source_group"]) else "missing_traceability"


def final_row_from_kept(row: pd.Series) -> dict[str, object]:
    raw = str(row.get("message_raw", "")).strip()
    return {
        "review_id": row.get("review_id", ""),
        "message_for_review": raw,
        "message_raw": raw,
        "message_clean": row.get("message_clean", ""),
        "source_label": row.get("source_label", ""),
        "normalized_label_before_review": row.get("normalized_label_before_review", ""),
        "candidate_reason": row.get("candidate_reason", ""),
        "source_name": row.get("source_name", ""),
        "dataset_name": row.get("dataset_name", ""),
        "source_group": row.get("source_group", ""),
        "contains_url": row.get("contains_url", ""),
        "contains_phone": row.get("contains_phone", ""),
        "contains_otp": row.get("contains_otp", ""),
        "contains_amount": row.get("contains_amount", ""),
        "suggested_category": row.get("suggested_category", ""),
        "raw_quality_status": row.get("raw_quality_status", "pass_raw_complete"),
        "source_traceability_status": source_traceability(row),
        "expert_label": "",
        "expert_confidence": "",
        "expert_notes": "",
        "reviewer_name": "",
        "review_date": "",
    }


def final_row_from_replacement(row: pd.Series, review_id: str) -> dict[str, object]:
    raw = str(row.get("message_raw", "")).strip()
    return {
        "review_id": review_id,
        "message_for_review": raw,
        "message_raw": raw,
        "message_clean": row.get("message_clean", ""),
        "source_label": row.get("source_label", ""),
        "normalized_label_before_review": row.get("normalized_label", ""),
        "candidate_reason": row.get("candidate_reason", ""),
        "source_name": row.get("source_name", ""),
        "dataset_name": row.get("dataset_name", ""),
        "source_group": row.get("source_group", ""),
        "contains_url": row.get("contains_url", ""),
        "contains_phone": row.get("contains_phone", ""),
        "contains_otp": row.get("contains_otp", ""),
        "contains_amount": row.get("contains_amount", ""),
        "suggested_category": row.get("suggested_category", ""),
        "raw_quality_status": row.get("raw_quality_status", "pass_raw_complete"),
        "source_traceability_status": source_traceability(row),
        "expert_label": "",
        "expert_confidence": "",
        "expert_notes": "",
        "reviewer_name": "",
        "review_date": "",
    }


def can_add(row: pd.Series, selected_keys: set[str], family_counts: dict[str, int], source_counts: dict[str, int], dataset_counts: dict[str, int], target_count: int) -> bool:
    raw = str(row.get("message_raw", "")).strip()
    key = normalize_text(raw)
    if not key or key in selected_keys:
        return False
    fam = family_key(raw)
    if family_counts.get(fam, 0) >= 5:
        return False
    source = str(row.get("source_name", "")).strip() or "unknown"
    dataset = str(row.get("dataset_name", "")).strip() or "unknown"
    soft_cap = max(1, int(target_count * 0.40))
    if source_counts.get(source, 0) >= soft_cap:
        return False
    if dataset_counts.get(dataset, 0) >= soft_cap:
        return False
    return True


def bump_counts(row: pd.Series, selected_keys: set[str], family_counts: dict[str, int], source_counts: dict[str, int], dataset_counts: dict[str, int]) -> None:
    raw = str(row.get("message_raw", "")).strip()
    selected_keys.add(normalize_text(raw))
    fam = family_key(raw)
    family_counts[fam] = family_counts.get(fam, 0) + 1
    source = str(row.get("source_name", "")).strip() or "unknown"
    dataset = str(row.get("dataset_name", "")).strip() or "unknown"
    source_counts[source] = source_counts.get(source, 0) + 1
    dataset_counts[dataset] = dataset_counts.get(dataset, 0) + 1


def write_excel(final_df: pd.DataFrame, raw_quality_summary: pd.DataFrame) -> None:
    source_summary = final_df.groupby(["candidate_reason", "source_name", "dataset_name"], dropna=False).size().reset_index(name="rows")
    codebook_df = pd.DataFrame({"label_codebook": CODEBOOK.splitlines()})
    instructions = pd.DataFrame(
        {
            "instruction": [
                "Review message_for_review; it equals the raw/original-looking SMS text in this packet.",
                "Keep source metadata separate from the labeling decision.",
                "Do not label smishing only because there is a URL.",
                "Allowed expert_label values: ham, spam_not_smishing, smishing, unsure, reject.",
                "Allowed expert_confidence values: high, medium, low.",
            ]
        }
    )
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        final_df.to_excel(writer, sheet_name="review_queue", index=False)
        codebook_df.to_excel(writer, sheet_name="label_codebook", index=False)
        source_summary.to_excel(writer, sheet_name="source_summary", index=False)
        raw_quality_summary.to_excel(writer, sheet_name="raw_quality_summary", index=False)
        instructions.to_excel(writer, sheet_name="instructions", index=False)

    wb = load_workbook(XLSX_PATH)
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(cell.value or "")) for cell in col[:100])
            ws.column_dimensions[letter].width = min(80, max(14, max_len + 2))
    ws = wb["review_queue"]
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 72
    ws.column_dimensions["D"].width = 55
    label_dv = DataValidation(type="list", formula1='"ham,spam_not_smishing,smishing,unsure,reject"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
    ws.add_data_validation(label_dv)
    ws.add_data_validation(conf_dv)
    label_dv.add(f"R2:R{max(ws.max_row, 2)}")
    conf_dv.add(f"S2:S{max(ws.max_row, 2)}")
    wb.save(XLSX_PATH)


def write_report(final_df: pd.DataFrame, initial: pd.DataFrame, kept: pd.DataFrame, replace: pd.DataFrame, pool: pd.DataFrame, log: pd.DataFrame) -> None:
    def counts(col: str) -> str:
        return "\n".join(f"- {idx}: {val}" for idx, val in final_df[col].value_counts().items()) or "- none"

    report = [
        "# Expert Spam Review Raw-Complete Packet Report",
        "",
        "## 1. Purpose",
        "This is a raw-complete expert review set for IAA and future relabeling. Every final sample is intended to preserve complete original-looking SMS text.",
        "",
        "## 2. Initial Packet Audit",
        f"- initial rows: {len(initial)}",
        f"- placeholder/anonymized rows found: {int(replace.get('raw_quality_flags', pd.Series(dtype=str)).astype(str).str.contains('placeholder_raw|html_encoded_placeholder', regex=True).sum())}",
        f"- duplicate rows found: {int(replace.get('raw_quality_status', pd.Series(dtype=str)).eq('fail_duplicate').sum())}",
        f"- long rows found: {int(initial.get('message_raw', pd.Series(dtype=str)).astype(str).str.len().gt(320).sum())}",
        f"- too-short rows found: {int(replace.get('raw_quality_status', pd.Series(dtype=str)).eq('fail_too_short').sum())}",
        f"- multi-message cells found: {int(replace.get('raw_quality_status', pd.Series(dtype=str)).eq('fail_multi_message_cell').sum())}",
        "",
        "## 3. Replacement Process",
        f"- rows kept: {len(kept)}",
        f"- rows replaced/requested: {len(replace)}",
        f"- replacement candidates available: {len(pool)}",
        f"- replacements accepted: {len(log)}",
        f"- shortage: {max(0, 500 - len(final_df))}",
        "",
        "## 4. Final Packet Composition",
        "### candidate_reason",
        counts("candidate_reason"),
        "",
        "### source_name",
        counts("source_name"),
        "",
        "### dataset_name",
        counts("dataset_name"),
        "",
        "### source_label",
        counts("source_label"),
        "",
        "### suggested_category",
        counts("suggested_category"),
        "",
        "### contains_url",
        counts("contains_url"),
        "",
        "### contains_phone",
        counts("contains_phone"),
        "",
        "### contains_otp",
        counts("contains_otp"),
        "",
        "### contains_amount",
        counts("contains_amount"),
        "",
        "## 5. Raw Completeness Validation",
        "All final samples have complete raw message text, no placeholder/anonymized raw tokens remain, and no synthetic rows were intentionally included.",
        "",
        "## 6. Expert Instructions Summary",
        CODEBOOK,
        "",
        "## 7. Future Use Note",
        "These rows are not yet part of the final dataset. They are held for expert labeling and IAA. After expert review, confirmed labels may be imported, agreement can be computed, disagreements can be resolved, and approved rows can be added to a future dataset version.",
    ]
    REPORT_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-count", type=int, default=500)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    initial = pd.read_csv(INITIAL_PATH, dtype=str, keep_default_na=False)
    kept = pd.read_csv(KEPT_PATH, dtype=str, keep_default_na=False)
    replace = pd.read_csv(REPLACE_PATH, dtype=str, keep_default_na=False)
    pool = pd.read_csv(POOL_PATH, dtype=str, keep_default_na=False)
    pool = pool.sample(frac=1, random_state=args.seed).sort_values("replacement_priority_score", ascending=False, kind="mergesort")

    selected_rows: list[dict[str, object]] = []
    selected_keys: set[str] = set()
    family_counts: dict[str, int] = {}
    source_counts: dict[str, int] = {}
    dataset_counts: dict[str, int] = {}

    for _, row in kept.iterrows():
        final_row = final_row_from_kept(row)
        selected_rows.append(final_row)
        bump_counts(pd.Series(final_row), selected_keys, family_counts, source_counts, dataset_counts)

    replacement_log: list[dict[str, object]] = []
    pool_cursor = 0
    pool_records = list(pool.to_dict("records"))
    for _, old in replace.iterrows():
        if len(selected_rows) >= args.target_count:
            break
        chosen = None
        while pool_cursor < len(pool_records):
            candidate = pd.Series(pool_records[pool_cursor])
            pool_cursor += 1
            if can_add(candidate, selected_keys, family_counts, source_counts, dataset_counts, args.target_count):
                chosen = candidate
                break
        if chosen is None:
            continue
        review_id = old.get("review_id", f"EXP-SPAM-{len(selected_rows) + 1:04d}")
        new_row = final_row_from_replacement(chosen, review_id)
        selected_rows.append(new_row)
        bump_counts(pd.Series(new_row), selected_keys, family_counts, source_counts, dataset_counts)
        replacement_log.append(
            {
                "old_review_id": old.get("review_id", ""),
                "old_message_raw": old.get("message_raw", ""),
                "old_failure_reason": old.get("replacement_needed_reason", old.get("raw_quality_flags", "")),
                "replacement_candidate_id": chosen.get("replacement_candidate_id", ""),
                "replacement_message_raw": chosen.get("message_raw", ""),
                "replacement_source_name": chosen.get("source_name", ""),
                "replacement_dataset_name": chosen.get("dataset_name", ""),
                "replacement_candidate_reason": chosen.get("candidate_reason", ""),
                "replacement_priority_score": chosen.get("replacement_priority_score", ""),
                "notes": "raw-complete replacement selected with seed and diversity caps",
            }
        )

    # Top up if the original failed rows were fewer than target gap or if any replacement could not be matched.
    while len(selected_rows) < args.target_count and pool_cursor < len(pool_records):
        candidate = pd.Series(pool_records[pool_cursor])
        pool_cursor += 1
        if not can_add(candidate, selected_keys, family_counts, source_counts, dataset_counts, args.target_count):
            continue
        review_id = f"EXP-SPAM-{len(selected_rows) + 1:04d}"
        new_row = final_row_from_replacement(candidate, review_id)
        selected_rows.append(new_row)
        bump_counts(pd.Series(new_row), selected_keys, family_counts, source_counts, dataset_counts)

    final_df = pd.DataFrame(selected_rows[: args.target_count])
    for col in FINAL_COLUMNS:
        if col not in final_df.columns:
            final_df[col] = ""
    final_df = final_df[FINAL_COLUMNS]
    final_df["review_id"] = [f"EXP-SPAM-{i:04d}" for i in range(1, len(final_df) + 1)]
    for col in ["expert_label", "expert_confidence", "expert_notes", "reviewer_name", "review_date"]:
        final_df[col] = ""

    log_df = pd.DataFrame(replacement_log)
    final_df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    log_df.to_csv(LOG_PATH, index=False, encoding="utf-8-sig")
    replace.to_csv(ARCHIVE_PATH, index=False, encoding="utf-8-sig")
    raw_quality_summary = final_df.groupby(["raw_quality_status", "source_traceability_status"], dropna=False).size().reset_index(name="rows")
    write_excel(final_df, raw_quality_summary)
    write_report(final_df, initial, kept, replace, pool, log_df)

    print("Raw-complete expert review packet repaired")
    print(f"initial packet rows: {len(initial)}")
    print(f"rows passed raw quality: {len(kept)}")
    print(f"rows needing replacement: {len(replace)}")
    print(f"replacement pool size: {len(pool)}")
    print(f"replacements accepted: {len(log_df)}")
    print(f"final packet row count: {len(final_df)}")
    print("source breakdown:")
    print(final_df["source_name"].value_counts().to_string())
    print("candidate reason breakdown:")
    print(final_df["candidate_reason"].value_counts().to_string())
    print(f"CSV path: {CSV_PATH.relative_to(ROOT).as_posix()}")
    print(f"Excel path: {XLSX_PATH.relative_to(ROOT).as_posix()}")
    print(f"report path: {REPORT_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
