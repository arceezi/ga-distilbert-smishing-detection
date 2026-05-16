#!/usr/bin/env python
"""Validate balanced raw-complete expert review packet."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
CSV_PATH = OUT_DIR / "expert_spam_review_500_balanced_raw_complete.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500_balanced_raw_complete.xlsx"
REPORT_PATH = OUT_DIR / "expert_spam_review_balanced_report.md"
CONV_PATH = OUT_DIR / "conversational_spam_candidate_pool.csv"

PLACEHOLDER_RE = re.compile(r"<\s*[A-Z0-9_ -]+\s*>")
URL_RE = re.compile(r"https?://\S+|www\.\S+|(?:[a-z0-9-]+\.)+[a-z]{2,}(?:/\S*)?", re.I)
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
AMOUNT_RE = re.compile(r"(?:[$£€]|rs\.?|php|usd|gbp|eur)\s*\d+(?:[,.]\d+)*|\d+(?:[,.]\d+)*(?:\s?(?:php|usd|gbp|eur|rs|p))", re.I)

REQUIRED = [
    "review_id", "message_for_review", "message_raw", "message_clean", "source_label",
    "normalized_label_before_review", "candidate_reason", "likely_review_bucket",
    "bucket_reason", "source_name", "dataset_name", "source_group", "contains_url",
    "contains_phone", "contains_otp", "contains_amount", "suggested_category",
    "raw_quality_status", "source_traceability_status", "expert_label",
    "expert_confidence", "expert_notes", "reviewer_name", "review_date",
]


def normalize_text(text: str) -> str:
    text = str(text).lower()
    text = URL_RE.sub(" <URL> ", text)
    text = re.sub(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", " <EMAIL> ", text)
    text = PHONE_RE.sub(" <PHONE> ", text)
    text = re.sub(r"\b(?:otp|pin|code|passcode|verification)\s*[:#-]?\s*[a-z0-9-]{4,10}\b", " <OTP> ", text)
    text = re.sub(r"\b[a-z]{1,4}\d{4,10}\b|\b\d{4,8}[a-z]{1,4}\b", " <OTP> ", text)
    text = re.sub(r"\b\d{9,}\b", " <REF_NUM> ", text)
    text = AMOUNT_RE.sub(" <AMOUNT> ", text)
    text = re.sub(r"[^a-z0-9<>]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def family_key(text: str) -> str:
    return " ".join(normalize_text(text).split()[:10])


def main() -> None:
    errors: list[str] = []
    warnings: list[str] = []
    df = pd.read_csv(CSV_PATH, dtype=str, keep_default_na=False)
    report = REPORT_PATH.read_text(encoding="utf-8") if REPORT_PATH.exists() else ""
    missing = [c for c in REQUIRED if c not in df.columns]
    if missing:
        errors.append(f"Missing columns: {missing}")
    if len(df) != 500 and "shortage:" not in report.lower():
        errors.append(f"Row count is {len(df)}, expected 500 unless shortage reported.")
    if "message_for_review" in df and "message_raw" in df:
        if not df["message_for_review"].equals(df["message_raw"]):
            errors.append("message_for_review != message_raw.")
        if df["message_raw"].str.strip().eq("").any():
            errors.append("Empty message_raw found.")
        if df["message_raw"].map(lambda x: bool(PLACEHOLDER_RE.search(str(x)) or "&lt;" in str(x).lower() or "&gt;" in str(x).lower())).any():
            errors.append("Placeholder/anonymized tokens found in message_raw.")
        if df["message_raw"].map(normalize_text).duplicated().any():
            errors.append("Duplicate normalized review messages found.")
        max_family = int(df["message_raw"].map(family_key).value_counts().max())
        if max_family > 5:
            errors.append(f"Template family cap exceeded: {max_family}.")
    if "is_synthetic" in df and df["is_synthetic"].astype(str).str.lower().isin(["true", "1", "yes"]).any():
        errors.append("Synthetic rows found.")
    for c in ["expert_label", "expert_confidence", "expert_notes", "reviewer_name", "review_date"]:
        if c in df.columns and not df[c].str.strip().eq("").all():
            errors.append(f"{c} is not blank.")
    if "source_label" in df and df["source_label"].str.strip().eq("").all():
        errors.append("Source labels not preserved.")
    if "source_name" in df:
        source_counts = df["source_name"].value_counts()
        if len(source_counts) and source_counts.iloc[0] > 200:
            errors.append(f"Source exceeds 40% cap: {source_counts.index[0]}={source_counts.iloc[0]}.")
        imc = int(source_counts.get("Smishing-Dataset-IMC25", 0))
        if imc > 200:
            errors.append(f"Smishing-Dataset-IMC25 exceeds 40%: {imc}.")
    if "likely_review_bucket" in df:
        buckets = df["likely_review_bucket"].value_counts()
        spam_count = int(buckets.get("likely_spam_not_smishing", 0))
        smish_count = int(buckets.get("likely_smishing", 0))
        unclear_count = int(buckets.get("unclear_review", 0))
        if not (260 <= smish_count <= 330 and spam_count >= 150 and unclear_count <= 60):
            if "shortage" not in report.lower():
                errors.append(f"Bucket counts outside target tolerance: smishing={smish_count}, spam={spam_count}, unclear={unclear_count}.")
        enough_spam = 0
        if CONV_PATH.exists():
            conv = pd.read_csv(CONV_PATH, dtype=str, keep_default_na=False)
            enough_spam = int(conv["likely_review_bucket"].eq("likely_spam_not_smishing").sum()) if "likely_review_bucket" in conv else 0
        if enough_spam >= 150 and spam_count < 150:
            errors.append(f"likely_spam_not_smishing count below 150 despite enough candidates: {spam_count}/{enough_spam}.")
    wb = load_workbook(XLSX_PATH, read_only=False)
    expected = {"review_queue", "label_codebook", "source_summary", "bucket_summary", "instructions"}
    if not expected.issubset(set(wb.sheetnames)):
        errors.append(f"Workbook missing sheets: {sorted(expected - set(wb.sheetnames))}")
    elif wb["review_queue"].max_row - 1 != len(df):
        errors.append("Workbook review_queue row count mismatch.")

    print("Balanced expert packet validation")
    print(f"final row count: {len(df)}")
    print("final likely bucket counts:")
    print(df["likely_review_bucket"].value_counts().to_string())
    print("final source breakdown:")
    print(df["source_name"].value_counts().to_string())
    if warnings:
        print("warnings:")
        for warning in warnings:
            print(f"- {warning}")
    if errors:
        print("validation failed:")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)
    print("validation passed: file is ready to send to expert")
    print(f"CSV path: {CSV_PATH.relative_to(ROOT).as_posix()}")
    print(f"Excel path: {XLSX_PATH.relative_to(ROOT).as_posix()}")
    print(f"report path: {REPORT_PATH.relative_to(ROOT).as_posix()}")
    print(f"most important file to send to expert: {XLSX_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
