#!/usr/bin/env python
"""Create the 500-row expert-review CSV/XLSX packet and documentation."""

from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
POOL_PATH = OUT_DIR / "expert_spam_review_source_pool.csv"
CSV_PATH = OUT_DIR / "expert_spam_review_500.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500.xlsx"
LOG_PATH = OUT_DIR / "expert_spam_review_sampling_log.csv"
CODEBOOK_PATH = OUT_DIR / "expert_spam_review_codebook.md"
REPORT_PATH = OUT_DIR / "expert_spam_review_report.md"
ARCHIVE_PATH = OUT_DIR / "expert_spam_review_excluded_archive.csv"

REVIEW_COLUMNS = [
    "review_id",
    "message_for_review",
    "message_raw",
    "message_clean",
    "source_label",
    "normalized_label_before_review",
    "candidate_reason",
    "source_name",
    "dataset_name",
    "source_group",
    "contains_url",
    "contains_phone",
    "contains_otp",
    "contains_amount",
    "suggested_category",
    "expert_label",
    "expert_confidence",
    "expert_notes",
    "reviewer_name",
    "review_date",
]

REASON_TARGETS = [
    ("original_spam_label", 220),
    ("needs_smishing_relabel", 80),
    ("possible_spam_not_smishing", 75),
    ("weak_signal_suspicious", 75),
    ("conflict_needs_review", 50),
    ("excluded_from_smishing_review", 50),
    ("public_candidate_spam", 50),
]

CODEBOOK = """# Expert Spam/Smishing Review Codebook

Purpose: label each SMS independently for expert review and inter-annotator agreement.

## HAM
Legitimate/non-malicious SMS.

Examples:
- normal OTP
- transaction alert
- delivery update
- telecom notice
- appointment reminder
- personal message

## SPAM_NOT_SMISHING
Unwanted promotional or irrelevant message but not clearly phishing.

Examples:
- generic ads
- gambling/casino/free-spin promos without credential theft
- adult/chat promo
- aggressive marketing
- random prize/reward promo without clear impersonation or credential/payment request

## SMISHING
SMS phishing/social-engineering attempt.

Examples:
- impersonates a bank, e-wallet, courier, telecom, government, or known service
- asks for login, OTP, password, PIN, payment, or account verification
- contains suspicious link/callback instruction
- threatens account lock/suspension
- uses financial/security/delivery urgency to make user act
- attempts credential theft or fraudulent payment

## UNSURE
Unclear, ambiguous, incomplete, or needs another reviewer.

## REJECT
Not useful for dataset.

Examples:
- not SMS-like
- OCR artifact
- non-English if out of scope
- abusive reply to scammer
- report/commentary text
- too incomplete
- duplicate fragment

Important: do not label a message as smishing only because it contains a URL. Look for deception, impersonation, credential/payment request, urgency, or social-engineering intent.

Allowed expert_label values: ham, spam_not_smishing, smishing, unsure, reject.

Allowed expert_confidence values: high, medium, low.
"""


def suggested_category(row: pd.Series) -> str:
    scam = str(row.get("scam_category", "")).strip()
    service = str(row.get("service_category", "")).strip()
    notes = str(row.get("notes", "")).lower()
    message = f"{row.get('message_raw', '')} {row.get('message_clean', '')}".lower()
    joined = f"{scam} {service} {notes} {message}".lower()
    checks = [
        ("banking/account-like suspicious", ["bank", "account", "wallet", "card", "verify", "suspend", "locked"]),
        ("delivery-like suspicious", ["delivery", "parcel", "package", "shipment", "courier", "usps", "ups", "fedex", "dhl"]),
        ("gambling/casino/free spin", ["casino", "gambling", "bet", "free spin", "jackpot", "my11circle"]),
        ("reward/prize", ["prize", "winner", "won", "reward", "bonus", "claim", "gift", "voucher"]),
        ("job/investment offer", ["job", "hiring", "income", "earn", "investment", "crypto", "forex"]),
        ("adult/chat promo", ["adult", "xxx", "sex", "chat", "dating", "porn"]),
        ("telecom promo", ["ringtone", "mobile", "txt", "airtime", "telecom", "call rate"]),
        ("promotional spam", ["free", "offer", "discount", "sale", "promo", "unsubscribe"]),
    ]
    for label, tokens in checks:
        if any(token in joined for token in tokens):
            return label
    return scam or service or "generic scam-like or unclear"


def source_family(row: pd.Series) -> str:
    key = str(row.get("expert_review_normalized_key", ""))
    words = key.split()
    return " ".join(words[:10]) if words else str(row.get("duplicate_cluster_id", ""))


def stratified_sample(pool: pd.DataFrame, target_count: int, seed: int) -> tuple[pd.DataFrame, pd.DataFrame]:
    valid = pool[pool["is_duplicate_representative"].astype(str).str.lower().eq("true")].copy()
    valid = valid[valid["message_raw"].fillna("").astype(str).str.strip().ne("") | valid["message_clean"].fillna("").astype(str).str.strip().ne("")]
    valid["_suggested_category"] = valid.apply(suggested_category, axis=1)
    valid["_source_family"] = valid.apply(source_family, axis=1)
    valid["_source_campaign"] = (
        valid["source_name"].fillna("").astype(str)
        + "||"
        + valid["_suggested_category"].fillna("").astype(str)
        + "||"
        + valid["_source_family"].fillna("").astype(str).str[:60]
    )
    valid = valid.sample(frac=1, random_state=seed).reset_index(drop=True)

    selected_idx: list[int] = []
    template_counts: dict[str, int] = {}
    campaign_counts: dict[str, int] = {}

    def try_add(frame: pd.DataFrame, quota: int) -> int:
        added = 0
        for idx, row in frame.iterrows():
            if idx in selected_idx:
                continue
            family = row["_source_family"]
            campaign = row["_source_campaign"]
            if template_counts.get(family, 0) >= 5:
                continue
            if campaign_counts.get(campaign, 0) >= 10:
                continue
            selected_idx.append(idx)
            template_counts[family] = template_counts.get(family, 0) + 1
            campaign_counts[campaign] = campaign_counts.get(campaign, 0) + 1
            added += 1
            if added >= quota or len(selected_idx) >= target_count:
                break
        return added

    log_rows = []
    remaining_target = target_count
    for reason, quota in REASON_TARGETS:
        if remaining_target <= 0:
            break
        reason_frame = valid[valid["candidate_reason"].eq(reason)]
        requested = min(quota, remaining_target)
        before = len(selected_idx)
        added = try_add(reason_frame, requested)
        remaining_target = target_count - len(selected_idx)
        log_rows.append(
            {
                "sampling_stage": "stratified_reason",
                "candidate_reason": reason,
                "available": len(reason_frame),
                "requested": requested,
                "selected": len(selected_idx) - before,
                "note": "underfilled if template/source caps or availability constrained selection",
            }
        )

    if len(selected_idx) < target_count:
        before = len(selected_idx)
        try_add(valid, target_count - len(selected_idx))
        log_rows.append(
            {
                "sampling_stage": "top_up",
                "candidate_reason": "any_valid_remaining",
                "available": len(valid) - before,
                "requested": target_count - before,
                "selected": len(selected_idx) - before,
                "note": "top-up used all candidate reasons while retaining caps",
            }
        )

    selected = valid.loc[selected_idx].copy()
    log_rows.append(
        {
            "sampling_stage": "final",
            "candidate_reason": "all",
            "available": len(valid),
            "requested": target_count,
            "selected": len(selected),
            "note": "shortage reported if selected is below requested target",
        }
    )
    return selected, pd.DataFrame(log_rows)


def build_review_frame(selected: pd.DataFrame) -> pd.DataFrame:
    selected = selected.reset_index(drop=True).copy()
    out = pd.DataFrame()
    out["review_id"] = [f"EXP-SPAM-{i:04d}" for i in range(1, len(selected) + 1)]
    out["message_for_review"] = selected["message_raw"].where(selected["message_raw"].fillna("").astype(str).str.strip().ne(""), selected["message_clean"])
    out["message_raw"] = selected["message_raw"]
    out["message_clean"] = selected["message_clean"]
    out["source_label"] = selected["source_label"]
    out["normalized_label_before_review"] = selected["normalized_label"]
    out["candidate_reason"] = selected["candidate_reason"]
    out["source_name"] = selected["source_name"]
    out["dataset_name"] = selected["dataset_name"]
    out["source_group"] = selected["source_group"]
    out["contains_url"] = selected["contains_url"]
    out["contains_phone"] = selected["contains_phone"]
    out["contains_otp"] = selected["contains_otp"]
    out["contains_amount"] = selected["contains_amount"]
    out["suggested_category"] = selected.apply(suggested_category, axis=1)
    for col in ["expert_label", "expert_confidence", "expert_notes", "reviewer_name", "review_date"]:
        out[col] = ""
    return out[REVIEW_COLUMNS]


def write_excel(review_df: pd.DataFrame, source_summary: pd.DataFrame) -> None:
    instructions = pd.DataFrame(
        {
            "instruction": [
                "Label only the SMS content in message_for_review.",
                "Use source metadata for audit only; do not let source labels override your judgment.",
                "Do not label smishing only because a message contains a URL.",
                "Use expert_label values: ham, spam_not_smishing, smishing, unsure, reject.",
                "Use expert_confidence values: high, medium, low.",
            ]
        }
    )
    codebook_rows = pd.DataFrame({"codebook": CODEBOOK.splitlines()})
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        review_df.to_excel(writer, sheet_name="review_queue", index=False)
        codebook_rows.to_excel(writer, sheet_name="label_codebook", index=False)
        source_summary.to_excel(writer, sheet_name="source_summary", index=False)
        instructions.to_excel(writer, sheet_name="instructions", index=False)

    wb = load_workbook(XLSX_PATH)
    ws = wb["review_queue"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    widths = {
        "A": 14,
        "B": 70,
        "C": 55,
        "D": 55,
        "E": 18,
        "F": 24,
        "G": 24,
        "H": 28,
        "I": 34,
        "J": 22,
        "O": 28,
        "P": 20,
        "Q": 20,
        "R": 36,
        "S": 22,
        "T": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    label_dv = DataValidation(type="list", formula1='"ham,spam_not_smishing,smishing,unsure,reject"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
    ws.add_data_validation(label_dv)
    ws.add_data_validation(conf_dv)
    label_dv.add(f"P2:P{max(ws.max_row, 2)}")
    conf_dv.add(f"Q2:Q{max(ws.max_row, 2)}")
    for sheet_name in ["label_codebook", "source_summary", "instructions"]:
        sheet = wb[sheet_name]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = min(80, max(18, max(len(str(c.value or "")) for c in col[:50]) + 2))
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(XLSX_PATH)


def write_report(review_df: pd.DataFrame, pool: pd.DataFrame, log_df: pd.DataFrame) -> None:
    reps = pool[pool["is_duplicate_representative"].astype(str).str.lower().eq("true")]
    dup_removed = int((pool["is_duplicate_representative"].astype(str).str.lower() != "true").sum())
    invalid_filtered = 0
    if ARCHIVE_PATH.exists():
        archive = pd.read_csv(ARCHIVE_PATH, dtype=str, keep_default_na=False)
        if "exclusion_type" in archive.columns:
            invalid_filtered = int(archive["exclusion_type"].astype(str).str.strip().ne("").sum())
    raw_candidates_found = len(pool) + invalid_filtered
    inspected = sorted(pool["source_file"].dropna().astype(str).unique())
    breakdowns = {
        "candidate_reason": review_df["candidate_reason"].value_counts(),
        "source_name": review_df["source_name"].value_counts(),
        "dataset_name": review_df["dataset_name"].value_counts(),
        "source_label": review_df["source_label"].value_counts(),
        "contains_url": review_df["contains_url"].value_counts(),
        "contains_otp": review_df["contains_otp"].value_counts(),
        "contains_phone": review_df["contains_phone"].value_counts(),
        "contains_amount": review_df["contains_amount"].value_counts(),
    }

    def md_counts(series: pd.Series) -> str:
        return "\n".join(f"- {idx}: {val}" for idx, val in series.items()) or "- none"

    report = [
        "# Expert Spam Review Packet Report",
        "",
        "## 1. Purpose",
        "This is a separate expert-review set for IAA/relabeling of spam or suspicious SMS messages. It is not part of the final training dataset.",
        "",
        "## 2. Source Files Inspected",
        *[f"- {item}" for item in inspected],
        "",
        "## 3. Candidate Pool Summary",
        f"- total raw candidates found: {raw_candidates_found}",
        f"- valid candidates after filtering: {len(pool)}",
        f"- representative candidates available after deduplication: {len(reps)}",
        f"- duplicates removed: {dup_removed}",
        f"- final review packet size: {len(review_df)}",
        "",
        "## 4. Sampling Strategy",
        "Sampling used seed 42 by default, reason-based strata, one representative per exact normalized duplicate cluster, a maximum of 5 per normalized template family, and a maximum of 10 per detectable broad source/campaign family.",
        "",
        "## 5. Review Packet Composition",
    ]
    for name, counts in breakdowns.items():
        report.extend([f"### {name}", md_counts(counts), ""])
    report.extend(
        [
            "## 6. Expert Codebook Summary",
            "Expert labels are ham, spam_not_smishing, smishing, unsure, and reject. Smishing requires deception, impersonation, credential/payment request, urgency, or social-engineering intent; a URL alone is not enough.",
            "",
            "## 7. Important Limitation",
            "This review packet is not yet part of the final training dataset. It is for expert review and IAA first.",
            "",
            "## 8. Next Step After Expert Review",
            "- import expert labels",
            "- compute agreement / IAA",
            "- resolve disagreements",
            "- create approved relabeled rows",
            "- decide whether to add confirmed smishing or ham/spam to future dataset versions",
        ]
    )
    REPORT_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-count", type=int, default=500)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    pool = pd.read_csv(POOL_PATH, dtype=str, keep_default_na=False)
    for col in ["is_duplicate_representative", "message_raw", "message_clean"]:
        if col not in pool.columns:
            raise SystemExit(f"Missing required pool column: {col}")

    selected, log_df = stratified_sample(pool, args.target_count, args.seed)
    review_df = build_review_frame(selected)
    source_summary = review_df.groupby(["candidate_reason", "source_name", "dataset_name"], dropna=False).size().reset_index(name="rows")

    review_df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    log_df.to_csv(LOG_PATH, index=False, encoding="utf-8-sig")
    CODEBOOK_PATH.write_text(CODEBOOK + f"\nGenerated: {date.today().isoformat()}\n", encoding="utf-8")
    write_excel(review_df, source_summary)
    write_report(review_df, pool, log_df)

    print("Expert spam review packet created")
    print(f"target row count: {args.target_count}")
    print(f"final review packet row count: {len(review_df)}")
    if len(review_df) < args.target_count:
        print(f"shortage: requested {args.target_count}, selected {len(review_df)}")
    print("review packet label/source breakdown:")
    print(review_df["candidate_reason"].value_counts().to_string())
    print(review_df["source_name"].value_counts().head(20).to_string())
    print(f"path to CSV: {CSV_PATH.relative_to(ROOT).as_posix()}")
    print(f"path to Excel: {XLSX_PATH.relative_to(ROOT).as_posix()}")
    print(f"path to codebook: {CODEBOOK_PATH.relative_to(ROOT).as_posix()}")
    print(f"path to report: {REPORT_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
