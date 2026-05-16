#!/usr/bin/env python
"""Validate the expert spam review CSV/XLSX packet."""

from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
CSV_PATH = OUT_DIR / "expert_spam_review_500.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500.xlsx"
POOL_PATH = OUT_DIR / "expert_spam_review_source_pool.csv"
LOG_PATH = OUT_DIR / "expert_spam_review_sampling_log.csv"
FINAL_V3_PATH = ROOT / "data" / "final_dataset_build" / "final" / "dataset_v3_public_manual_research_synthetic_ham_balanced.csv"

REQUIRED_COLUMNS = [
    "review_id",
    "message_for_review",
    "message_raw",
    "message_clean",
    "source_label",
    "normalized_label_before_review",
    "candidate_reason",
    "source_name",
    "dataset_name",
    "source_group",
    "contains_url",
    "contains_phone",
    "contains_otp",
    "contains_amount",
    "suggested_category",
    "expert_label",
    "expert_confidence",
    "expert_notes",
    "reviewer_name",
    "review_date",
]


def normalize_review_text(text: str) -> str:
    text = str(text).lower()
    text = re.sub(r"https?://\S+|www\.\S+|(?:[a-z0-9-]+\.)+[a-z]{2,}(?:/\S*)?", " <URL> ", text)
    text = re.sub(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", " <EMAIL> ", text)
    text = re.sub(r"(?:\+?\d[\d\s().-]{6,}\d)", " <PHONE> ", text)
    text = re.sub(r"\b(?:otp|pin|code|passcode|verification)\s*[:#-]?\s*[a-z0-9-]{4,10}\b", " <OTP> ", text)
    text = re.sub(r"\b[a-z]{1,4}\d{4,10}\b|\b\d{4,8}[a-z]{1,4}\b", " <OTP> ", text)
    text = re.sub(r"\b\d{9,}\b", " <REF_NUM> ", text)
    text = re.sub(r"(?:[$£€]|rs\.?|php|usd|gbp|eur)\s*\d+(?:[,.]\d+)*|\d+(?:[,.]\d+)*(?:\s?(?:php|usd|gbp|eur|rs|p))", " <AMOUNT> ", text)
    text = re.sub(r"[^a-z0-9<>]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def source_family(text: str) -> str:
    return " ".join(normalize_review_text(text).split()[:10])


def fail(errors: list[str], message: str) -> None:
    errors.append(message)


def main() -> None:
    errors: list[str] = []
    warnings: list[str] = []
    if not CSV_PATH.exists():
        raise SystemExit(f"Missing CSV: {CSV_PATH}")
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing XLSX: {XLSX_PATH}")
    if not POOL_PATH.exists():
        raise SystemExit(f"Missing source pool: {POOL_PATH}")

    review = pd.read_csv(CSV_PATH, dtype=str, keep_default_na=False)
    pool = pd.read_csv(POOL_PATH, dtype=str, keep_default_na=False)
    log = pd.read_csv(LOG_PATH, dtype=str, keep_default_na=False) if LOG_PATH.exists() else pd.DataFrame()

    target = 500
    shortage_reported = False
    if not log.empty:
        final_selected = int(log.iloc[-1].get("selected", len(review)))
        shortage_reported = final_selected < target
    if len(review) != target and not shortage_reported:
        fail(errors, f"Target row count is {target}, found {len(review)}, and no shortage was reported.")

    missing = [col for col in REQUIRED_COLUMNS if col not in review.columns]
    if missing:
        fail(errors, f"Missing required review columns: {missing}")
    else:
        if review["message_for_review"].str.strip().eq("").any():
            fail(errors, "Empty message_for_review values found.")
        keys = review["message_for_review"].map(normalize_review_text)
        if keys.duplicated().any():
            fail(errors, "Exact duplicate normalized review messages found.")
        families = review["message_for_review"].map(source_family)
        max_family = int(families.value_counts().max()) if len(families) else 0
        if max_family > 5:
            fail(errors, f"Campaign/template family cap exceeded: max family count {max_family}.")
        for col in ["expert_label", "expert_confidence", "expert_notes", "reviewer_name", "review_date"]:
            if not review[col].str.strip().eq("").all():
                fail(errors, f"{col} is not blank before expert review.")
        if review["source_label"].str.strip().eq("").all():
            fail(errors, "Source labels were not preserved.")
        trace_cols = ["source_name", "dataset_name", "source_group"]
        no_trace = review[trace_cols].fillna("").astype(str).apply(lambda r: not any(v.strip() for v in r), axis=1)
        if no_trace.any():
            fail(errors, "Rows without source traceability found.")

    if "is_synthetic" in pool.columns and pool["is_synthetic"].astype(str).str.lower().isin(["true", "1", "yes"]).any():
        fail(errors, "Synthetic rows found in source pool.")
    if FINAL_V3_PATH.exists():
        final = pd.read_csv(FINAL_V3_PATH, dtype=str, keep_default_na=False)
        synth_ham = final[
            final.get("is_synthetic", "").astype(str).str.lower().isin(["true", "1", "yes"])
            & final.get("normalized_label", "").astype(str).str.lower().eq("ham")
        ]
        synth_keys = set(synth_ham["message_raw"].where(synth_ham["message_raw"].str.strip().ne(""), synth_ham["message_clean"]).map(normalize_review_text))
        overlap = set(review["message_for_review"].map(normalize_review_text)) & synth_keys
        if overlap:
            fail(errors, f"Final V3 synthetic ham overlap found: {len(overlap)} normalized messages.")

    wb = load_workbook(XLSX_PATH, read_only=False)
    expected_sheets = {"review_queue", "label_codebook", "source_summary", "instructions"}
    if not expected_sheets.issubset(set(wb.sheetnames)):
        fail(errors, f"Workbook missing expected sheets: {sorted(expected_sheets - set(wb.sheetnames))}")
    else:
        ws = wb["review_queue"]
        if ws.max_row - 1 != len(review):
            fail(errors, f"XLSX review_queue row count {ws.max_row - 1} does not match CSV {len(review)}.")
        if ws.freeze_panes != "A2":
            warnings.append("XLSX review_queue top row is not frozen at A2.")
        if not ws.auto_filter.ref:
            warnings.append("XLSX review_queue filters are not enabled.")

    print("Expert spam review validation")
    print(f"CSV rows: {len(review)}")
    print(f"XLSX sheets: {', '.join(wb.sheetnames)}")
    print(f"source pool rows: {len(pool)}")
    print(f"required expert columns: present" if not missing else f"required expert columns missing: {missing}")
    if warnings:
        print("warnings:")
        for warning in warnings:
            print(f"- {warning}")
    if errors:
        print("validation failed:")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)
    print("validation passed: file is suitable to send to expert")
    print(f"most important file to send to expert: {XLSX_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
