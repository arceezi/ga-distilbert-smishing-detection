#!/usr/bin/env python
"""Validate the raw-complete expert review packet."""

from __future__ import annotations

import ast
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
CSV_PATH = OUT_DIR / "expert_spam_review_500_raw_complete.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500_raw_complete.xlsx"
REPORT_PATH = OUT_DIR / "expert_spam_review_raw_complete_report.md"
FINAL_V3_PATH = ROOT / "data" / "final_dataset_build" / "final" / "dataset_v3_public_manual_research_synthetic_ham_balanced.csv"

PLACEHOLDER_RE = re.compile(r"<\s*[A-Z0-9_ -]+\s*>")
URL_RE = re.compile(r"https?://\S+|www\.\S+|(?:[a-z0-9-]+\.)+[a-z]{2,}(?:/\S*)?", re.I)
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
AMOUNT_RE = re.compile(r"(?:[$£€]|rs\.?|php|usd|gbp|eur)\s*\d+(?:[,.]\d+)*|\d+(?:[,.]\d+)*(?:\s?(?:php|usd|gbp|eur|rs|p))", re.I)

REQUIRED_COLUMNS = [
    "review_id",
    "message_for_review",
    "message_raw",
    "message_clean",
    "source_label",
    "normalized_label_before_review",
    "candidate_reason",
    "source_name",
    "dataset_name",
    "source_group",
    "contains_url",
    "contains_phone",
    "contains_otp",
    "contains_amount",
    "suggested_category",
    "raw_quality_status",
    "source_traceability_status",
    "expert_label",
    "expert_confidence",
    "expert_notes",
    "reviewer_name",
    "review_date",
]


def normalize_text(text: str) -> str:
    text = str(text).lower()
    text = URL_RE.sub(" <URL> ", text)
    text = re.sub(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", " <EMAIL> ", text)
    text = PHONE_RE.sub(" <PHONE> ", text)
    text = re.sub(r"\b(?:otp|pin|code|passcode|verification)\s*[:#-]?\s*[a-z0-9-]{4,10}\b", " <OTP> ", text)
    text = re.sub(r"\b[a-z]{1,4}\d{4,10}\b|\b\d{4,8}[a-z]{1,4}\b", " <OTP> ", text)
    text = re.sub(r"\b\d{9,}\b", " <REF_NUM> ", text)
    text = AMOUNT_RE.sub(" <AMOUNT> ", text)
    text = re.sub(r"[^a-z0-9<>]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def family_key(text: str) -> str:
    return " ".join(normalize_text(text).split()[:10])


def multi_message_cell(text: str) -> bool:
    raw = str(text).strip()
    if raw.startswith("[") and raw.endswith("]"):
        try:
            parsed = ast.literal_eval(raw)
            return isinstance(parsed, list) and len(parsed) > 1
        except Exception:
            return True
    return raw.count("', '") >= 1 or raw.count('", "') >= 1


def artifact_like(text: str) -> bool:
    raw = str(text)
    low = raw.lower()
    if any(term in low for term in ["screenshot", "ocr", "image may contain", "reported by", "commentary", "reply to scammer"]):
        return True
    if raw.count("|") >= 4 or raw.count("\t") >= 2:
        return True
    if re.search(r"\b(row|label|source|dataset)\s*[:=]", low) and len(raw) > 120:
        return True
    return False


def main() -> None:
    errors: list[str] = []
    warnings: list[str] = []
    if not CSV_PATH.exists():
        raise SystemExit(f"Missing CSV: {CSV_PATH}")
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing XLSX: {XLSX_PATH}")
    df = pd.read_csv(CSV_PATH, dtype=str, keep_default_na=False)
    report_text = REPORT_PATH.read_text(encoding="utf-8") if REPORT_PATH.exists() else ""

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        errors.append(f"Missing columns: {missing}")
    if len(df) != 500 and "shortage:" not in report_text.lower():
        errors.append(f"Target count is 500, found {len(df)}, and no shortage reported.")
    if "is_synthetic" in df.columns and df["is_synthetic"].astype(str).str.lower().isin(["true", "1", "yes"]).any():
        errors.append("Synthetic rows found.")
    if "message_for_review" in df.columns and "message_raw" in df.columns:
        if df["message_for_review"].str.strip().eq("").any():
            errors.append("Empty message_for_review values found.")
        if not df["message_for_review"].equals(df["message_raw"]):
            errors.append("message_for_review does not equal message_raw.")
        if df["message_raw"].map(lambda x: bool(PLACEHOLDER_RE.search(str(x)) or "&lt;" in str(x).lower() or "&gt;" in str(x).lower())).any():
            errors.append("Angle-bracket placeholder/anonymized tokens remain in message_raw.")
        if df["message_raw"].map(multi_message_cell).any():
            errors.append("Multi-message list/cell rows found.")
        if df["message_raw"].map(artifact_like).any():
            errors.append("UI/OCR/report artifacts found.")
        keys = df["message_raw"].map(normalize_text)
        if keys.duplicated().any():
            errors.append("Exact duplicate normalized review messages found.")
        max_family = int(df["message_raw"].map(family_key).value_counts().max()) if len(df) else 0
        if max_family > 5:
            errors.append(f"Campaign/template family cap exceeded: {max_family}.")
    if "raw_text_status" in df.columns and df["raw_text_status"].astype(str).str.lower().eq("already_redacted").any():
        errors.append("raw_text_status == already_redacted found.")
    if "raw_text_available" in df.columns and df["raw_text_available"].astype(str).str.lower().isin(["false", "0", "no"]).any():
        errors.append("raw_text_available == False found.")
    for col in ["expert_label", "expert_confidence", "expert_notes", "reviewer_name", "review_date"]:
        if col in df.columns and not df[col].str.strip().eq("").all():
            errors.append(f"{col} is not blank.")
    if "source_label" in df.columns and df["source_label"].str.strip().eq("").all():
        errors.append("Source labels are not preserved.")
    if "source_traceability_status" in df.columns and df["source_traceability_status"].ne("traceable").any():
        errors.append("Missing source traceability found.")
    if "source_name" in df.columns and len(df):
        max_source = df["source_name"].value_counts().iloc[0]
        if max_source > len(df) * 0.40:
            msg = f"source cap exceeded: {max_source}/{len(df)} from one source"
            if "unless unavoidable" in report_text.lower() or "source cap" in report_text.lower():
                warnings.append(msg)
            else:
                errors.append(msg)
    if "dataset_name" in df.columns and len(df):
        max_dataset = df["dataset_name"].value_counts().iloc[0]
        if max_dataset > len(df) * 0.40:
            msg = f"dataset cap exceeded: {max_dataset}/{len(df)} from one dataset"
            if "unless unavoidable" in report_text.lower() or "dataset cap" in report_text.lower():
                warnings.append(msg)
            else:
                errors.append(msg)

    if FINAL_V3_PATH.exists() and "message_raw" in df.columns:
        final = pd.read_csv(FINAL_V3_PATH, dtype=str, keep_default_na=False)
        synth_ham = final[
            final.get("is_synthetic", pd.Series("", index=final.index)).astype(str).str.lower().isin(["true", "1", "yes"])
            & final.get("normalized_label", pd.Series("", index=final.index)).astype(str).str.lower().eq("ham")
        ]
        synth_keys = set(synth_ham["message_raw"].where(synth_ham["message_raw"].str.strip().ne(""), synth_ham["message_clean"]).map(normalize_text))
        overlap = set(df["message_raw"].map(normalize_text)) & synth_keys
        if overlap:
            errors.append(f"Final V3 synthetic ham overlap found: {len(overlap)}.")

    wb = load_workbook(XLSX_PATH, read_only=False)
    expected = {"review_queue", "label_codebook", "source_summary", "raw_quality_summary", "instructions"}
    if not expected.issubset(set(wb.sheetnames)):
        errors.append(f"Missing workbook sheets: {sorted(expected - set(wb.sheetnames))}")
    else:
        ws = wb["review_queue"]
        if ws.max_row - 1 != len(df):
            errors.append("Workbook review_queue row count does not match CSV.")
        if ws.freeze_panes != "A2":
            warnings.append("Workbook review_queue top row is not frozen at A2.")
        if not ws.auto_filter.ref:
            warnings.append("Workbook review_queue filters are not enabled.")

    placeholder_remaining = 0
    duplicate_remaining = 0
    if "message_raw" in df.columns:
        placeholder_remaining = int(df["message_raw"].map(lambda x: bool(PLACEHOLDER_RE.search(str(x)) or "&lt;" in str(x).lower() or "&gt;" in str(x).lower())).sum())
        duplicate_remaining = int(df["message_raw"].map(normalize_text).duplicated().sum())

    print("Raw-complete expert packet validation")
    print(f"final packet row count: {len(df)}")
    print(f"placeholder/anonymized rows remaining: {placeholder_remaining}")
    print(f"duplicate rows remaining: {duplicate_remaining}")
    if "source_name" in df.columns:
        print("source breakdown:")
        print(df["source_name"].value_counts().to_string())
    if "candidate_reason" in df.columns:
        print("candidate reason breakdown:")
        print(df["candidate_reason"].value_counts().to_string())
    if warnings:
        print("warnings:")
        for warning in warnings:
            print(f"- {warning}")
    if errors:
        print("validation failed:")
        for error in errors:
            print(f"- {error}")
        raise SystemExit(1)
    print("validation passed: file is suitable to send to expert and may be used after expert review")
    print(f"CSV path: {CSV_PATH.relative_to(ROOT).as_posix()}")
    print(f"Excel path: {XLSX_PATH.relative_to(ROOT).as_posix()}")
    print(f"report path: {REPORT_PATH.relative_to(ROOT).as_posix()}")
    print(f"most important file to send to expert: {XLSX_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
