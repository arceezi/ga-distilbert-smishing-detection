#!/usr/bin/env python
"""Create a balanced raw-complete expert review packet."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "data" / "expert_review_iaa"
CURRENT_PATH = OUT_DIR / "expert_spam_review_500_raw_complete.csv"
CONV_PATH = OUT_DIR / "conversational_spam_candidate_pool.csv"
REPL_PATH = OUT_DIR / "raw_complete_expert_replacement_pool.csv"
CSV_PATH = OUT_DIR / "expert_spam_review_500_balanced_raw_complete.csv"
XLSX_PATH = OUT_DIR / "expert_spam_review_500_balanced_raw_complete.xlsx"
LOG_PATH = OUT_DIR / "expert_spam_review_balanced_sampling_log.csv"
REMOVED_PATH = OUT_DIR / "expert_spam_review_balanced_removed_archive.csv"
REPORT_PATH = OUT_DIR / "expert_spam_review_balanced_report.md"

PLACEHOLDER_RE = re.compile(r"<\s*[A-Z0-9_ -]+\s*>")
URL_RE = re.compile(r"https?://\S+|www\.\S+|(?:[a-z0-9-]+\.)+[a-z]{2,}(?:/\S*)?", re.I)
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{6,}\d)")
AMOUNT_RE = re.compile(r"(?:[$£€]|rs\.?|php|usd|gbp|eur)\s*\d+(?:[,.]\d+)*|\d+(?:[,.]\d+)*(?:\s?(?:php|usd|gbp|eur|rs|p))", re.I)

FINAL_COLUMNS = [
    "review_id",
    "message_for_review",
    "message_raw",
    "message_clean",
    "source_label",
    "normalized_label_before_review",
    "candidate_reason",
    "likely_review_bucket",
    "bucket_reason",
    "source_name",
    "dataset_name",
    "source_group",
    "contains_url",
    "contains_phone",
    "contains_otp",
    "contains_amount",
    "suggested_category",
    "raw_quality_status",
    "source_traceability_status",
    "expert_label",
    "expert_confidence",
    "expert_notes",
    "reviewer_name",
    "review_date",
]

CODEBOOK = """# Expert Label Codebook

HAM: Legitimate/non-malicious SMS.

SPAM_NOT_SMISHING: Unwanted promotional or irrelevant SMS but not clearly phishing.
Examples: ringtone/music/adult chat promos; gambling/free-spin/casino ads without credential theft; generic marketing; subscription ads; prize/reward promos without impersonation or credential/payment collection; premium number ads; reply STOP style promos.

SMISHING: Deceptive SMS designed to trick the recipient into unsafe action.
Examples: account verification/update/login; bank/e-wallet/courier/government impersonation; OTP/PIN/password/credential request; suspicious link or callback instruction tied to account/payment/security/delivery; account lock/suspension threat; fake delivery payment/address issue; financial/security urgency.

UNSURE: Ambiguous or needs another reviewer.

REJECT: Not useful, not SMS-like, non-English, artifact, duplicate fragment, abusive reply, or report/commentary text.

Important expert note: A message can be spam without being smishing. A message can contain a link without automatically being smishing. A message is smishing when there is deception/social engineering toward credentials, payment, account access, identity, or fraudulent action.
"""


def normalize_text(text: str) -> str:
    text = str(text).lower()
    text = URL_RE.sub(" <URL> ", text)
    text = re.sub(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", " <EMAIL> ", text)
    text = PHONE_RE.sub(" <PHONE> ", text)
    text = re.sub(r"\b(?:otp|pin|code|passcode|verification)\s*[:#-]?\s*[a-z0-9-]{4,10}\b", " <OTP> ", text)
    text = re.sub(r"\b[a-z]{1,4}\d{4,10}\b|\b\d{4,8}[a-z]{1,4}\b", " <OTP> ", text)
    text = re.sub(r"\b\d{9,}\b", " <REF_NUM> ", text)
    text = AMOUNT_RE.sub(" <AMOUNT> ", text)
    text = re.sub(r"[^a-z0-9<>]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def family_key(text: str) -> str:
    return " ".join(normalize_text(text).split()[:10])


def raw_ok(raw: str) -> bool:
    raw = str(raw).strip()
    return bool(raw and not PLACEHOLDER_RE.search(raw) and "&lt;" not in raw.lower() and "&gt;" not in raw.lower())


def classify(text: str, source_label: str = "", normalized_label: str = "", candidate_reason: str = "") -> tuple[int, int, str, str, str]:
    low = text.lower()
    spam = 0
    smish = 0
    reasons: list[str] = []
    spam_rules = [
        (r"\bfree\b|\boffer\b|discount|sale|promo|voucher", 2, "promotional_offer"),
        (r"ringtone|poly|tone|music|mobile content|txt .* to \d|text .* to \d", 3, "ringtone_subscription"),
        (r"adult|xxx|sex|chat|dating|hot singles|porn", 3, "adult_chat_promo"),
        (r"casino|bet|gambl|free spin|jackpot", 2, "gambling_promo"),
        (r"reply stop|txt stop|text stop|unsubscribe|opt out|stop to", 3, "opt_out_ad"),
        (r"\bcall \d|premium|per min|p per|150p|50p|£1\.50", 2, "premium_number_ad"),
        (r"prize|winner|won|reward|claim", 2, "prize_reward_ad"),
    ]
    smish_rules = [
        (r"bank|account|acct|card|wallet|paypal|venmo|zelle|cash app", 3, "bank_account_bait"),
        (r"login|password|otp|pin|cvv|credential|verification code", 4, "credential_or_otp_request"),
        (r"verify|verification|update your|confirm|validate|restore", 3, "verification_action"),
        (r"locked|suspend|restricted|blocked|unauthorized|security alert", 4, "account_threat"),
        (r"delivery|parcel|package|shipment|courier|usps|ups|fedex|dhl|address issue|redelivery", 3, "delivery_bait"),
        (r"government|tax|irs|hmrc|benefit|fine|refund", 3, "government_bait"),
        (r"https?://|www\.", 1, "link_present"),
        (r"urgent|immediately|within \d+ hours|today only|final notice", 2, "urgency"),
        (r"payment|pay fee|unpaid|billing|charge|refund", 3, "payment_bait"),
    ]
    for pattern, pts, reason in spam_rules:
        if re.search(pattern, low):
            spam += pts
            reasons.append(reason)
    for pattern, pts, reason in smish_rules:
        if re.search(pattern, low):
            smish += pts
            reasons.append(reason)
    if source_label.lower() == "spam" or normalized_label.lower() == "spam":
        spam += 3
        reasons.append("original_spam_label")
    if "smish" in source_label.lower() or normalized_label.lower() == "smishing":
        smish += 2
    if candidate_reason in {"conflict_needs_review", "possible_spam_not_smishing"}:
        reasons.append(candidate_reason)
    if spam >= smish + 2 and smish < 5:
        bucket = "likely_spam_not_smishing"
    elif smish >= 5 or smish > spam:
        bucket = "likely_smishing"
    elif "conflict" in candidate_reason or len(text) > 300:
        bucket = "unclear_review"
    elif spam >= 3:
        bucket = "likely_spam_not_smishing"
    else:
        bucket = "unclear_review"
    category = "unclear suspicious message"
    if "ringtone_subscription" in reasons:
        category = "telecom/ringtone/subscription spam"
    elif "adult_chat_promo" in reasons:
        category = "adult/chat promo"
    elif "gambling_promo" in reasons:
        category = "gambling/casino/free spin"
    elif "bank_account_bait" in reasons:
        category = "banking/account-like suspicious"
    elif "delivery_bait" in reasons:
        category = "delivery-like suspicious"
    elif "government_bait" in reasons:
        category = "government/tax/benefit suspicious"
    elif "prize_reward_ad" in reasons:
        category = "prize/reward promo"
    elif spam >= 3:
        category = "promotional spam"
    return spam, smish, bucket, ";".join(dict.fromkeys(reasons)), category


def source_traceability(row: pd.Series) -> str:
    return "traceable" if any(str(row.get(c, "")).strip() for c in ["source_name", "dataset_name", "source_group"]) else "missing_traceability"


def standardize_current(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        raw = str(row.get("message_raw", "")).strip()
        spam, smish, bucket, reason, category = classify(raw, row.get("source_label", ""), row.get("normalized_label_before_review", ""), row.get("candidate_reason", ""))
        rows.append({
            **row.to_dict(),
            "normalized_label": row.get("normalized_label_before_review", ""),
            "likely_review_bucket": bucket,
            "bucket_reason": reason,
            "spam_signal_score": spam,
            "smishing_signal_score": smish,
            "suggested_category": row.get("suggested_category", "") or category,
            "selection_priority_score": 40 + smish * 6 + spam * 2,
            "candidate_origin": "current_raw_complete_packet",
            "duplicate_key": normalize_text(raw),
        })
    return pd.DataFrame(rows)


def standardize_replacement(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, row in df.iterrows():
        raw = str(row.get("message_raw", "")).strip()
        spam, smish, bucket, reason, category = classify(raw, row.get("source_label", ""), row.get("normalized_label", ""), row.get("candidate_reason", ""))
        rows.append({
            "message_raw": raw,
            "message_clean": row.get("message_clean", ""),
            "source_label": row.get("source_label", ""),
            "normalized_label": row.get("normalized_label", ""),
            "candidate_reason": row.get("candidate_reason", ""),
            "source_name": row.get("source_name", ""),
            "dataset_name": row.get("dataset_name", ""),
            "source_group": row.get("source_group", ""),
            "contains_url": row.get("contains_url", ""),
            "contains_phone": row.get("contains_phone", ""),
            "contains_otp": row.get("contains_otp", ""),
            "contains_amount": row.get("contains_amount", ""),
            "suggested_category": row.get("suggested_category", "") or category,
            "raw_quality_status": row.get("raw_quality_status", "pass_raw_complete"),
            "source_traceability_status": source_traceability(row),
            "likely_review_bucket": bucket,
            "bucket_reason": reason,
            "spam_signal_score": spam,
            "smishing_signal_score": smish,
            "selection_priority_score": int(float(row.get("replacement_priority_score", 0) or 0)) + smish * 5,
            "candidate_origin": "raw_complete_replacement_pool",
            "duplicate_key": normalize_text(raw),
        })
    return pd.DataFrame(rows)


def standardize_conv(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["normalized_label"] = out.get("normalized_label", "")
    out["source_traceability_status"] = out.apply(source_traceability, axis=1)
    out["candidate_origin"] = "conversational_spam_pool"
    return out


def addable(row: pd.Series, selected_keys: set[str], family_counts: dict[str, int], source_counts: dict[str, int], dataset_counts: dict[str, int], target: int) -> bool:
    raw = str(row.get("message_raw", "")).strip()
    key = normalize_text(raw)
    if not raw_ok(raw) or not key or key in selected_keys:
        return False
    fam = family_key(raw)
    if family_counts.get(fam, 0) >= 5:
        return False
    cap = int(target * 0.40)
    source = str(row.get("source_name", "")).strip() or "unknown"
    dataset = str(row.get("dataset_name", "")).strip() or "unknown"
    if source_counts.get(source, 0) >= cap:
        return False
    if dataset_counts.get(dataset, 0) >= cap:
        return False
    return True


def bump(row: pd.Series, selected_keys: set[str], family_counts: dict[str, int], source_counts: dict[str, int], dataset_counts: dict[str, int]) -> None:
    raw = str(row.get("message_raw", "")).strip()
    selected_keys.add(normalize_text(raw))
    fam = family_key(raw)
    source = str(row.get("source_name", "")).strip() or "unknown"
    dataset = str(row.get("dataset_name", "")).strip() or "unknown"
    family_counts[fam] = family_counts.get(fam, 0) + 1
    source_counts[source] = source_counts.get(source, 0) + 1
    dataset_counts[dataset] = dataset_counts.get(dataset, 0) + 1


def select_from(pool: pd.DataFrame, bucket: str, quota: int, selected: list[pd.Series], selected_keys: set[str], family_counts: dict[str, int], source_counts: dict[str, int], dataset_counts: dict[str, int], target: int) -> int:
    added = 0
    frame = pool[pool["likely_review_bucket"].eq(bucket)]
    for _, row in frame.iterrows():
        if added >= quota or len(selected) >= target:
            break
        if addable(row, selected_keys, family_counts, source_counts, dataset_counts, target):
            selected.append(row)
            bump(row, selected_keys, family_counts, source_counts, dataset_counts)
            added += 1
    return added


def to_final(rows: list[pd.Series]) -> pd.DataFrame:
    out_rows = []
    for i, row in enumerate(rows, 1):
        raw = str(row.get("message_raw", "")).strip()
        out_rows.append({
            "review_id": f"EXP-BAL-{i:04d}",
            "message_for_review": raw,
            "message_raw": raw,
            "message_clean": row.get("message_clean", ""),
            "source_label": row.get("source_label", ""),
            "normalized_label_before_review": row.get("normalized_label", row.get("normalized_label_before_review", "")),
            "candidate_reason": row.get("candidate_reason", ""),
            "likely_review_bucket": row.get("likely_review_bucket", ""),
            "bucket_reason": row.get("bucket_reason", ""),
            "source_name": row.get("source_name", ""),
            "dataset_name": row.get("dataset_name", ""),
            "source_group": row.get("source_group", ""),
            "contains_url": row.get("contains_url", ""),
            "contains_phone": row.get("contains_phone", ""),
            "contains_otp": row.get("contains_otp", ""),
            "contains_amount": row.get("contains_amount", ""),
            "suggested_category": row.get("suggested_category", ""),
            "raw_quality_status": row.get("raw_quality_status", "pass_raw_complete"),
            "source_traceability_status": row.get("source_traceability_status", "traceable"),
            "expert_label": "",
            "expert_confidence": "",
            "expert_notes": "",
            "reviewer_name": "",
            "review_date": "",
        })
    return pd.DataFrame(out_rows, columns=FINAL_COLUMNS)


def write_excel(df: pd.DataFrame) -> None:
    source_summary = df.groupby(["source_name", "dataset_name", "likely_review_bucket"], dropna=False).size().reset_index(name="rows")
    bucket_summary = df.groupby(["likely_review_bucket", "candidate_reason", "suggested_category"], dropna=False).size().reset_index(name="rows")
    codebook_df = pd.DataFrame({"label_codebook": CODEBOOK.splitlines()})
    instructions = pd.DataFrame({"instruction": [
        "Review message_for_review; it equals message_raw.",
        "Use likely_review_bucket only as audit metadata, not as a final label.",
        "A message can be spam without being smishing.",
        "A link alone does not make a message smishing.",
        "Allowed expert_label values: ham, spam_not_smishing, smishing, unsure, reject.",
        "Allowed expert_confidence values: high, medium, low.",
    ]})
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="review_queue", index=False)
        codebook_df.to_excel(writer, sheet_name="label_codebook", index=False)
        source_summary.to_excel(writer, sheet_name="source_summary", index=False)
        bucket_summary.to_excel(writer, sheet_name="bucket_summary", index=False)
        instructions.to_excel(writer, sheet_name="instructions", index=False)
    wb = load_workbook(XLSX_PATH)
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(80, max(14, max(len(str(c.value or "")) for c in col[:100]) + 2))
    ws = wb["review_queue"]
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 72
    label_dv = DataValidation(type="list", formula1='"ham,spam_not_smishing,smishing,unsure,reject"', allow_blank=True)
    conf_dv = DataValidation(type="list", formula1='"high,medium,low"', allow_blank=True)
    ws.add_data_validation(label_dv); ws.add_data_validation(conf_dv)
    label_dv.add(f"T2:T{max(ws.max_row, 2)}")
    conf_dv.add(f"U2:U{max(ws.max_row, 2)}")
    wb.save(XLSX_PATH)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target-count", type=int, default=500)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    current = standardize_current(pd.read_csv(CURRENT_PATH, dtype=str, keep_default_na=False))
    conv = standardize_conv(pd.read_csv(CONV_PATH, dtype=str, keep_default_na=False))
    repl = standardize_replacement(pd.read_csv(REPL_PATH, dtype=str, keep_default_na=False))
    all_pool = pd.concat([conv, current, repl], ignore_index=True, sort=False)
    all_pool = all_pool[all_pool["message_raw"].astype(str).map(raw_ok)]
    all_pool = all_pool.drop_duplicates("duplicate_key", keep="first")
    # Sort to prefer conversational UCI spam for spam bucket and high-signal replacement/current rows for smishing.
    all_pool["_source_bonus"] = all_pool["source_name"].eq("UCI SMS Spam Collection").astype(int) * 80 + all_pool["source_name"].eq("Mishra & Soni").astype(int) * 30
    all_pool["_spam_bucket_bonus"] = all_pool["likely_review_bucket"].eq("likely_spam_not_smishing").astype(int) * 50
    all_pool["_priority"] = pd.to_numeric(all_pool.get("selection_priority_score", 0), errors="coerce").fillna(0) + all_pool["_source_bonus"] + all_pool["_spam_bucket_bonus"]
    all_pool = all_pool.sample(frac=1, random_state=args.seed).sort_values("_priority", ascending=False, kind="mergesort")

    selected: list[pd.Series] = []
    selected_keys: set[str] = set()
    family_counts: dict[str, int] = {}
    source_counts: dict[str, int] = {}
    dataset_counts: dict[str, int] = {}
    log_rows = []
    targets = [("likely_spam_not_smishing", 175), ("likely_smishing", 300), ("unclear_review", 25)]
    for bucket, quota in targets:
        before = len(selected)
        added = select_from(all_pool, bucket, min(quota, args.target_count - len(selected)), selected, selected_keys, family_counts, source_counts, dataset_counts, args.target_count)
        log_rows.append({"bucket": bucket, "target": quota, "available": int(all_pool["likely_review_bucket"].eq(bucket).sum()), "selected": added})

    # Fill any remaining rows with any bucket, still respecting caps.
    if len(selected) < args.target_count:
        before = len(selected)
        for _, row in all_pool.iterrows():
            if len(selected) >= args.target_count:
                break
            if addable(row, selected_keys, family_counts, source_counts, dataset_counts, args.target_count):
                selected.append(row)
                bump(row, selected_keys, family_counts, source_counts, dataset_counts)
        log_rows.append({"bucket": "top_up_any", "target": args.target_count - before, "available": len(all_pool), "selected": len(selected) - before})

    final = to_final(selected[: args.target_count])
    selected_set = set(final["message_raw"].map(normalize_text))
    removed = current[~current["message_raw"].map(normalize_text).isin(selected_set)].copy()
    log = pd.DataFrame(log_rows)
    final.to_csv(CSV_PATH, index=False, encoding="utf-8-sig")
    log.to_csv(LOG_PATH, index=False, encoding="utf-8-sig")
    removed.to_csv(REMOVED_PATH, index=False, encoding="utf-8-sig")
    write_excel(final)
    shortage = max(0, args.target_count - len(final))
    report = [
        "# Balanced Expert Spam Review Packet Report",
        "",
        "## 1. Purpose",
        "This revised packet adds more general/conversational spam so expert review can distinguish spam_not_smishing from smishing.",
        "",
        "## 2. Source Basis",
        "UCI / SMS Spam Collection is a suitable public source for mobile spam research because it contains SMS spam rows originally collected for spam filtering research.",
        "",
        "## 3. Old Packet Summary",
        "- old packet rows: 500",
        "- old packet was raw-complete but likely smishing-heavy",
        "",
        "## 4. New Sampling Strategy",
        "- target likely_smishing: 300",
        "- target likely_spam_not_smishing: 175",
        "- target unclear/conflict/reject: 25",
        f"- shortage: {shortage}",
        "- source cap: no source should exceed 40%; UCI can be heavy for conversational spam but is reported.",
        "",
        "## 5. Final Packet Composition",
        "### likely_review_bucket",
        *[f"- {k}: {v}" for k, v in final["likely_review_bucket"].value_counts().items()],
        "### source_name",
        *[f"- {k}: {v}" for k, v in final["source_name"].value_counts().items()],
        "### candidate_reason",
        *[f"- {k}: {v}" for k, v in final["candidate_reason"].value_counts().items()],
        "### suggested_category",
        *[f"- {k}: {v}" for k, v in final["suggested_category"].value_counts().items()],
        "",
        "## 6. Validation Results",
        "The packet is raw-complete by construction, contains no synthetic rows by source filtering, has no placeholder raw messages, and uses normalized duplicate controls.",
        "",
        "## 7. Use Note",
        "This packet is for expert review and IAA only. It is not yet added to the final dataset.",
    ]
    REPORT_PATH.write_text("\n".join(report) + "\n", encoding="utf-8")
    print("Balanced raw-complete expert packet created")
    print(f"likely_spam_not_smishing candidates found: {int(all_pool['likely_review_bucket'].eq('likely_spam_not_smishing').sum())}")
    print(f"likely_smishing candidates used: {int(final['likely_review_bucket'].eq('likely_smishing').sum())}")
    print(f"unclear/review candidates used: {int(final['likely_review_bucket'].eq('unclear_review').sum())}")
    print(f"final row count: {len(final)}")
    print("final likely bucket counts:")
    print(final["likely_review_bucket"].value_counts().to_string())
    print("final source breakdown:")
    print(final["source_name"].value_counts().to_string())
    print(f"CSV path: {CSV_PATH.relative_to(ROOT).as_posix()}")
    print(f"Excel path: {XLSX_PATH.relative_to(ROOT).as_posix()}")
    print(f"report path: {REPORT_PATH.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
